from flask import Flask, request
import pandas as pd
from openpyxl import load_workbook
import os

app = Flask(__name__)

EXCEL_FILE = "event_data.xlsx"

@app.route("/submit", methods=["POST"])
def submit():
    data = request.json
    school = data.get("school")
    staff1 = data.get("staff1")
    staff2 = data.get("staff2")

    rows = []
    for i in range(1, 26):
        student = data.get(f"student{i}")
        if student:
            events = []
            for event in ["Dance", "Singing", "Drawing", "Essay", "Quiz", "Drama", "Sports"]:
                key = f"event_{i}_{event}"
                events.append(1 if key in data else 0)
            rows.append([school, staff1, staff2, student] + events)

    df = pd.DataFrame(rows, columns=["School", "Staff 1", "Staff 2", "Student",
                                     "Dance", "Singing", "Drawing", "Essay", "Quiz", "Drama", "Sports"])
    if os.path.exists(EXCEL_FILE):
        book = load_workbook(EXCEL_FILE)
        with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl', mode='a', if_sheet_exists='overlay') as writer:
            writer.book = book
            df.to_excel(writer, index=False, header=False, startrow=book.active.max_row)
    else:
        df.to_excel(EXCEL_FILE, index=False)

    return "Data Saved Successfully!"
